import openpyxl

print('loading workbook...')
wb = openpyxl.load_workbook('automate_online-materials/produceSales.xlsx')
sheet = wb['Sheet']

for i in range(1,len(sheet['A'])+1):
    print(f'{sheet.cell(row=i, column=1).value} : {sheet.cell(row=i, column=2).value}')
    if sheet.cell(row=i,column=1).value == 'Celery':
        sheet['B'+str(i)] = 1.19

    if sheet.cell(row=i,column=1).value == 'Garlic':
        sheet['B'+str(i)] = 3.07

    if sheet.cell(row=i,column=1).value == 'Lemon':
        sheet['B'+str(i)] = 1.27

wb.save('/Users/chanlim/Desktop/Updated_PS.xlsx')