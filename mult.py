import openpyxl
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Font

wb = openpyxl.Workbook()

st = wb['Sheet']

st['A1'] = '곱셈표'

for i in range(2, 11):
    st['A'+str(i)] = i-1
    st['A'+str(i)].font = Font(bold=True)

for i in range(2, 11):
    st[get_column_letter(i)+'1'] = i-1
    st[get_column_letter(i)+'1'].font = Font(bold=True)

for i in range(1,10):
    for j in range(1,10):
        st[get_column_letter(i+1)+str(j+1)] = i*j

wb.save('/Users/chanlim/Desktop/multiplication_chart.xlsx')



